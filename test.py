from cgitb import text
from playwright.sync_api import sync_playwright
from creds import username, password
from bs4 import BeautifulSoup as bs
import openpyxl as xl

wb = xl.load_workbook('Expenses.xlsx')
ws = wb['2nd Year']


months = ['January', 
          'February', 
          'March', 
          'April', 
          'May', 
          'June', 
          'July', 
          'August', 
          'September', 
          'October', 
          'November', 
          'December']

def Pull_Data():
    with sync_playwright() as p:
        browser = p.chromium.launch()
        #browser = p.chromium.launch(headless=False, slow_mo=50)
        page = browser.new_page()
        page.goto('https://www.appalachianpower.com/account/login/')
        page.fill('input#cphContentMain_ctl00_ctl00_TbUserID', username)
        page.fill('input#cphContentMain_ctl00_ctl00_TbPassword', password)
        page.click('input[type=submit]')
        html1 = page.inner_html('#cphContentMain_ctl00_DivSummary')
        html2 = page.inner_html('#cphContentMain_ctl00_ctl06_RptPaymentList_TrHistoryRow_0')
        soup1 = bs(html1,'html.parser')
        soup2 = bs(html2,'html.parser')

        due_now = soup1.find('div', {'class': 'd-flex'}).text
        Pull_Data.due_now = due_now[1:6] #This is the amount due this month
        due_previous = soup2.find_all('td')
        Pull_Data.month_previous = int(str(due_previous[0])[13:15]) #this is the number associated to last month, ex: 05 for May
        Pull_Data.due_previous = str(due_previous[2])[5:10] #this is the amount due last month

def Update():
    for index, month in enumerate(months,start=1):
        cell_months = ws.cell(index+1, 1)
        if cell_months.value == str(months[Pull_Data.month_previous-1]):
            cell_previous = ws.cell(index+1,3)
            cell_int_previous = ws.cell(index+1,4)
            cell_previous.value = float(Pull_Data.due_previous)/2
            cell_now = ws.cell(index+2,3)
            cell_int_now = ws.cell(index+1,4)
            cell_now.value = float(Pull_Data.due_now)/2
        else:
            pass

    wb.save('Expenses.xlsx')