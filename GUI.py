from asyncore import write
import tkinter as tk
import test
from cgitb import text
from playwright.sync_api import sync_playwright
from creds import username, password
from bs4 import BeautifulSoup as bs
import openpyxl as xl

root = tk.Tk()
root.geometry("250x150")
# test.Pull_Data()
# test.Update()

wb = xl.load_workbook('Expenses.xlsx', data_only=True)
ws = wb['2nd Year']

months = ['January','February','March','April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

with sync_playwright() as p:
    browser = p.chromium.launch()
    #browser = p.chromium.launch(headless=False, slow_mo=50)
    page = browser.new_page()
    page.goto('https://www.appalachianpower.com/account/login/')
    page.fill('input#cphContentMain_ctl00_ctl00_TbUserID', username)
    page.fill('input#cphContentMain_ctl00_ctl00_TbPassword', password)
    page.click('input[type=submit]')
    html1 = page.inner_html('#cphContentMain_ctl00_DivSummary')
    html2 = page.inner_html('#cphContentMain_ctl00_ctl06_RptPaymentList_TrHistoryRow_0')
    soup1 = bs(html1,'html.parser')
    soup2 = bs(html2,'html.parser')

    due_now = soup1.find('div', {'class': 'd-flex'}).text
    due_now = due_now[1:6] #This is the amount due this month
    due_prev = soup2.find_all('td')
    month_prev = int(str(due_prev[0])[13:15]) #this is the number associated to last month, ex: 05 for May
    due_prev = str(due_prev[2])[5:10] #this is the amount due last month

for index, month in enumerate(months,start=1):
    cell_months = ws.cell(index+1, 1)
    if cell_months.value == str(months[month_prev-1]):
        cell_prev = ws.cell(index+1,3)
        cell_prev.value = float(due_prev)/2
        cell_int_prev = ws.cell(index+1,4).value
        cell_tot_prev = 0
        for num in range(3,7):
            cell_tot_prev += int(ws.cell(index+1,num).value)

        cell_now = ws.cell(index+2,3)
        cell_now.value = float(due_now)/2
        cell_int_now = ws.cell(index+2,4).value
        cell_tot_now = 0
        for num in range(3,7):
            cell_tot_now += int(ws.cell(index+2,num).value)
        break
    else:
        pass

wb.save('Expenses.xlsx')

prev = tk.Label(root, text=str(months[month_prev-1]))
prev.grid(row=1,column=3)
now = tk.Label(root, text=str(months[month_prev]))
now.grid(row=1,column=4)

elec = tk.Label(root, text = "Electricity")
elec.grid(row=2,column=2)
elec_data_prev = tk.Label(root,text=f'${float(due_prev)/2}')
elec_data_prev.grid(row=2,column=3)
elec_data_now = tk.Label(root,text=f'${float(due_now)/2}')
elec_data_now.grid(row=2,column=4)

int = tk.Label(root, text = "Internet")
int_data_prev= tk.Label(root,text=f'${round(cell_int_prev,2)}')
int_data_prev.grid(row=3,column=3)
int_data_now= tk.Label(root,text=f'${round(cell_int_now,2)}')
int.grid(row=3,column=2)
int_data_now.grid(row=3,column=4)

laundry = tk.Label(root, text = "Laundry")
laundry.grid(row=4,column=2)
other = tk.Label(root, text = "Other")
other.grid(row=5,column=2)

def Total():
    tot=tk.Label(root,text="Total")
    tot.grid(row=6,column=2)
    tot_prev=tk.Label(root,text=f'${round(cell_tot_prev,2)}')
    tot_prev.grid(row=6,column=3)
    tot_now=tk.Label(root,text=f'${round(cell_tot_now,2)}')
    tot_now.grid(row=6,column=4)

Total()

laundry_prev = tk.Entry(root,width=10)
laundry_prev.grid(row=4,column=3)
laundry_now = tk.Entry(root,width=10)
laundry_now.grid(row=4,column=4)

other_prev = tk.Entry(root,width=10)
other_prev.grid(row=5,column=3)
other_now = tk.Entry(root,width=10)
other_now.grid(row=5,column=4)

def Input():
    ws.cell(index+1,5,int(laundry_prev.get()))
    ws.cell(index+2,5,int(laundry_now.get()))
    ws.cell(index+1,6,int(other_prev.get()))
    ws.cell(index+2,6,int(other_now.get()))
    wb.save('Expenses.xlsx')

    cell_tot_prev = 0
    for num in range(3,7):
        cell_tot_prev += int(ws.cell(index+1,num).value)

    cell_tot_now = 0
    for num in range(3,7):
        cell_tot_now += int(ws.cell(index+2,num).value)
    
    Total()


wb.save('Expenses.xlsx')

update = tk.Button(root, text="Update", command=Input)
update.grid(row=7, column=3)

root.mainloop()


